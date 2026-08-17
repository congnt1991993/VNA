import openpyxl

file_path = "/Users/congnguyen/Library/CloudStorage/OneDrive-Personal/Work/04. Company/Synodus/GOV/01.VNA/Document/Chỉ tiêu Governance_v3_Edited.xlsx"

wb = openpyxl.load_workbook(file_path)
sheet = wb.active # or wb['Chi tiết chỉ tiêu']
print("Active sheet:", sheet.title)
print("Dimensions:", sheet.dimensions)

for row in range(1, 4):
    row_vals = [sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)]
    print(f"Row {row}: {row_vals}")

print("Merged cells:")
for r in sheet.merged_cells.ranges:
    print(r)
