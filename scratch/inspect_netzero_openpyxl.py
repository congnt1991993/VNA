import openpyxl

file_path = "/Users/congnguyen/Library/CloudStorage/OneDrive-Personal/Work/04. Company/Synodus/GOV/01.VNA/Document/Netzero_GRI Summary.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb['95 Chỉ tiêu']

print("Row 1 cells:")
print([sheet.cell(1, c).value for c in range(1, 20)])
print("\nRow 2 cells:")
print([sheet.cell(2, c).value for c in range(1, 20)])
print("\nRow 3 cells:")
print([sheet.cell(3, c).value for c in range(1, 20)])

print("\nMerged cells:")
for r in list(sheet.merged_cells.ranges)[:10]:
    print(r)

print("\nFirst 5 data rows:")
for r in range(1, 10):
    row_vals = [sheet.cell(r, c).value for c in range(1, 15)]
    print(f"Row {r}: {row_vals}")
