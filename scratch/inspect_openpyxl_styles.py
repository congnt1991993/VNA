import openpyxl

file_path = "/Users/congnguyen/Library/CloudStorage/OneDrive-Personal/Work/04. Company/Synodus/GOV/01.VNA/Document/Chỉ tiêu Governance_v3_Edited.xlsx"

wb = openpyxl.load_workbook(file_path)
sheet = wb['Chi tiết chỉ tiêu']

# Inspect styles of column I (Index 9) and header cells
for row in [1, 3, 4]:
    cell_i = sheet.cell(row, 9)
    print(f"Row {row} Col I value: {cell_i.value}")
    print(f"Row {row} Col I font: {cell_i.font.name if cell_i.font else None}, size: {cell_i.font.size if cell_i.font else None}, bold: {cell_i.font.bold if cell_i.font else None}, color: {cell_i.font.color.rgb if cell_i.font and cell_i.font.color else None}")
    print(f"Row {row} Col I fill: {cell_i.fill.fill_type if cell_i.fill else None}, fgColor: {cell_i.fill.fgColor.rgb if cell_i.fill and cell_i.fill.fgColor else None}")
    print(f"Row {row} Col I alignment: horizontal={cell_i.alignment.horizontal if cell_i.alignment else None}, vertical={cell_i.alignment.vertical if cell_i.alignment else None}, wrap_text={cell_i.alignment.wrap_text if cell_i.alignment else None}")
    print(f"Row {row} Col I border: {cell_i.border}")
    print("-" * 40)

# Print column widths
for col_idx in range(1, 10):
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    print(f"Column {col_letter} width: {sheet.column_dimensions[col_letter].width}")
