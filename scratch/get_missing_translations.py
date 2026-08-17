import openpyxl

file_path = "/Users/congnguyen/Library/CloudStorage/OneDrive-Personal/Work/04. Company/Synodus/GOV/01.VNA/Document/Netzero_GRI Summary.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb['95 Chỉ tiêu']

print("Row index | GRI Code | Description")
print("-" * 80)
for r in range(3, sheet.max_row + 1):
    kpi_name = sheet.cell(r, 13).value
    desc = sheet.cell(r, 7).value
    if kpi_name is None:
        print(f"Row {r:2d} | {sheet.cell(r, 4).value} | {repr(desc)}")
