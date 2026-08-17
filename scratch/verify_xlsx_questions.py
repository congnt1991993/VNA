import openpyxl

file_path = "/Users/congnguyen/Library/CloudStorage/OneDrive-Personal/Work/04. Company/Synodus/GOV/01.VNA/Document/Chỉ tiêu Governance_v3_Edited.xlsx"

wb = openpyxl.load_workbook(file_path)
sheet = wb['Chi tiết chỉ tiêu']

print("Row index | STT | GRI Code | J cell value")
print("-" * 80)
for r in range(3, sheet.max_row + 1):
    stt = sheet.cell(r, 1).value
    gri = sheet.cell(r, 2).value
    val = sheet.cell(r, 10).value
    short_val = (val[:60] + "...") if val and len(val) > 60 else val
    print(f"Row {r:2d} | {stt} | {gri} | {short_val}")

print("\nTitle row merged range check:")
for r in sheet.merged_cells.ranges:
    if r.min_row == 1:
        print(f"Merged range in Row 1: {r.coord}")

print("\nCol J Width:", sheet.column_dimensions['J'].width)
