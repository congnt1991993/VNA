import openpyxl

file_path = "/Users/congnguyen/Library/CloudStorage/OneDrive-Personal/Work/04. Company/Synodus/GOV/01.VNA/Document/Chỉ tiêu Governance_v3_Edited.xlsx"

wb = openpyxl.load_workbook(file_path)
sheet = wb['Chi tiết chỉ tiêu']

print("Row index | STT | GRI Code | Indicator (Vie)")
print("-" * 50)
for r in range(4, sheet.max_row + 1):
    stt = sheet.cell(r, 1).value
    gri = sheet.cell(r, 2).value
    vie = sheet.cell(r, 4).value
    print(f"Row {r:2d} | {stt} | {gri} | {vie}")
