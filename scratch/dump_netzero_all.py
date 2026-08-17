import openpyxl

file_path = "/Users/congnguyen/Library/CloudStorage/OneDrive-Personal/Work/04. Company/Synodus/GOV/01.VNA/Document/Netzero_GRI Summary.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb['95 Chỉ tiêu']

with open("/Users/congnguyen/Library/CloudStorage/OneDrive-Personal/Work/04. Company/Synodus/GOV/01.VNA/VNA/scratch/inspect_netzero_all_rows_output.txt", "w", encoding="utf-8") as f:
    f.write(f"Max row: {sheet.max_row}\n\n")
    for r in range(3, sheet.max_row + 1):
        desc = sheet.cell(r, 7).value
        kpi_name = sheet.cell(r, 13).value
        gri_code = sheet.cell(r, 4).value
        f.write(f"Row {r:2d} | GRI: {gri_code}\n")
        f.write(f"       | Col G (Description): {desc}\n")
        f.write(f"       | Col M (Tên KPI):      {kpi_name}\n")
        f.write("-" * 80 + "\n")
print("Done writing inspect_netzero_all_rows_output.txt")
