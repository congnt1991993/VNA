import openpyxl

file_path = "/Users/congnguyen/Library/CloudStorage/OneDrive-Personal/Work/04. Company/Synodus/GOV/01.VNA/Document/Netzero_GRI Summary.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb['95 Chỉ tiêu']

print(f"Max row: {sheet.max_row}")
for r in range(3, sheet.max_row + 1):
    desc = sheet.cell(r, 7).value
    kpi_name = sheet.cell(r, 13).value
    print(f"Row {r:2d} | Col G (Description): {repr(desc)}")
    print(f"       | Col M (Tên KPI):      {repr(kpi_name)}")
    print("-" * 50)
