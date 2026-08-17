import os
import shutil
import openpyxl

# File paths
file_path = "/Users/congnguyen/Library/CloudStorage/OneDrive-Personal/Work/04. Company/Synodus/GOV/01.VNA/Document/Netzero_GRI Summary.xlsx"
backup_dir = "/Users/congnguyen/Library/CloudStorage/OneDrive-Personal/Work/04. Company/Synodus/GOV/01.VNA/Document/Backup"
backup_path = os.path.join(backup_dir, "Netzero_GRI Summary_Backup.xlsx")

print("--- Start Netzero Translations Update ---")

# Step 1: Create Backup
print("Checking backup directory...")
if not os.path.exists(backup_dir):
    os.makedirs(backup_dir)
    print("Created backup directory:", backup_dir)

print(f"Creating backup of original file to: {backup_path}")
shutil.copy(file_path, backup_path)
print("Backup created successfully.")

# Step 2: Load workbook
print(f"Loading workbook from: {file_path}")
wb = openpyxl.load_workbook(file_path)
sheet = wb['95 Chỉ tiêu']

# Row-based translation mapping
translation_map = {
    4: "- Kiểm kê khí nhà kính các chuyến bay nội địa (Phát thải khí nhà kính trực tiếp (Phạm vi 1) cho các chuyến bay nội địa)\n- Lượng nhiên liệu tra nạp từ các sân bay trong nước (chuyến bay nội địa, hoặc quốc tế có điểm xuất phát từ Việt Nam)",
    5: "Nhận diện mối nguy, đánh giá rủi ro và điều tra sự cố",
    8: "Thỏa ước lao động tập thể",
    9: "Nghĩa vụ chương trình phúc lợi xác định và các chương trình hưu trí khác",
    10: "Tỷ lệ lương khởi điểm chuẩn theo giới tính so với mức lương tối thiểu vùng",
    11: "Tỷ lệ nhân sự quản lý cấp cao được tuyển dụng từ cộng đồng địa phương",
    13: "Phúc lợi dành cho nhân viên toàn thời gian mà không áp dụng cho nhân viên tạm thời hoặc bán thời gian",
    14: "Số ngày làm việc bị mất do đình công",
    15: "Sự tham gia, tham vấn và truyền thông của người lao động về an toàn và sức khỏe nghề nghiệp",
    16: "Tai nạn lao động liên quan đến công việc",
    17: "Bệnh nghề nghiệp liên quan đến công việc",
    19: "Các vụ việc phân biệt đối xử và biện pháp khắc phục đã thực hiện",
    20: "Các chương trình nâng cao kỹ năng cho nhân viên và các chương trình hỗ trợ chuyển đổi",
    21: "Tỷ lệ nhân viên nhận được đánh giá định kỳ về hiệu quả công việc và phát triển nghề nghiệp",
    22: "Sự hài lòng của nhân viên",
    24: "Đánh giá tác động sức khỏe và an toàn của các nhóm sản phẩm và dịch vụ",
    30: "Tỷ lệ chi tiêu cho các nhà cung cấp địa phương",
    37: "Báo cáo phát thải CO2 (Báo cáo lượng CO2 phát thải quá mức cho phép và phải được bù trừ bằng tín chỉ CO2 (EUA))",
    38: "Quản lý các chủ đề trọng yếu",
    39: "Các cam kết về chính sách",
    40: "Cơ chế tham vấn và phản ánh các vấn đề đạo đức",
    41: "Cơ cấu và thành phần quản trị",
    42: "Ủy quyền trách nhiệm",
    43: "Vai trò của cơ quan quản trị cao nhất trong giám sát quản lý tác động",
    44: "Chủ tịch cơ quan quản trị cao nhất",
    45: "Quy trình đề cử và lựa chọn cơ quan quản lý cao nhất",
    46: "Xung đột lợi ích",
    47: "Tiếp cận đối thoại các bên liên quan",
    48: "Hỗ trợ tài chính nhận được từ chính phủ",
    49: "Truyền thông và đào tạo về các chính sách và quy trình chống tham nhũng",
    50: "Sự cố tham nhũng đã xác nhận và biện pháp xử lý",
    51: "Các vụ kiện pháp lý về hành vi phản cạnh tranh, chống độc quyền và độc quyền",
    52: "Đóng góp chính trị",
    56: "Khí oxit nitơ (NOx), oxit lưu huỳnh (SOx) và các khí thải độc hại đáng kể khác vào không khí",
    57: "Phát thải các chất làm suy giảm tầng ôzôn (ODS)",
    58: "Các sự cố tràn đổ đáng kể",
    59: "Quản lý các chủ đề trọng yếu",
    60: "Các hoạt động có tác động tiêu cực đáng kể (thực tế và tiềm ẩn) đến cộng đồng địa phương",
    61: "Tỷ lệ tổng thu nhập hàng năm",
    62: "Quản lý các chủ đề trọng yếu",
    63: "Các hoạt động, chuỗi giá trị và các mối quan hệ kinh doanh khác",
    64: "Nghỉ phép cha mẹ (thai sản/nuôi con)",
    65: "Thời gian thông báo tối thiểu về các thay đổi vận hành",
    66: "Tỷ lệ lương cơ bản và thù lao giữa lao động nữ và nam",
    67: "Các hoạt động và nhà cung cấp có rủi ro đối với quyền tự do liên kết và thỏa ước lao động tập thể",
    68: "Các hoạt động và nhà cung cấp có rủi ro lớn về sử dụng lao động trẻ em",
    69: "Các hoạt động và nhà cung cấp có rủi ro lớn về lao động cưỡng bức hoặc bắt buộc",
    70: "Các hoạt động đã được rà soát hoặc đánh giá tác động về quyền con người",
    71: "Các nhà cung cấp mới được sàng lọc theo các tiêu chí xã hội",
    72: "Tác động tiêu cực về mặt xã hội trong chuỗi cung ứng và biện pháp đã thực hiện",
    73: "Quản lý các chủ đề trọng yếu",
    74: "An toàn và sức khỏe nghề nghiệp",
    75: "Quản lý các chủ đề trọng yếu",
    76: "Yêu cầu về thông tin sản phẩm, dịch vụ và nhãn mác",
    77: "Tương tác với nguồn nước như một tài nguyên dùng chung",
    78: "Quản lý các tác động liên quan đến xả nước",
    79: "Xả nước",
    80: "Phát thải khí nhà kính gián tiếp do năng lượng (Phạm vi 2)",
    81: "Quản lý các chủ đề trọng yếu",
    82: "Các hoạt động, chuỗi giá trị và các mối quan hệ kinh doanh khác",
    83: "Tác động tiêu cực về môi trường trong chuỗi cung ứng và các hành động được thực hiện",
    84: "Tác động tiêu cực về mặt xã hội trong chuỗi cung ứng và biện pháp đã thực hiện",
    85: "Quản lý các chủ đề trọng yếu",
    86: "Nguyên vật liệu được sử dụng theo trọng lượng hoặc thể tích",
    87: "Nguyên vật liệu đầu vào tái chế được sử dụng",
    88: "Sản phẩm và vật liệu bao bì được thu hồi",
    89: "Phát sinh chất thải và tác động đáng kể liên quan đến chất thải",
    90: "Chất thải phát sinh",
    91: "Chất thải được chuyển hướng khỏi nơi thải bỏ",
    92: "Chất thải được chuyển đến nơi thải bỏ",
    93: "Tác động tiêu cực về môi trường trong chuỗi cung ứng và các hành động được thực hiện",
    94: "Quản lý các chủ đề trọng yếu",
    95: "Tiêu thụ năng lượng bên ngoài tổ chức",
    96: "Phát thải khí nhà kính gián tiếp khác (Phạm vi 3)",
    97: "CARBURE (Báo cáo sử dụng nhiên liệu sinh học bền vững tại Pháp)"
}

# Step 3: Write translations to Column M
print("Writing translations to Column M...")
for r in range(3, sheet.max_row + 1):
    kpi_cell = sheet.cell(row=r, column=13) # Column M
    desc_cell = sheet.cell(row=r, column=7)  # Column G
    
    if kpi_cell.value is None or str(kpi_cell.value).strip() == "":
        if r in translation_map:
            kpi_cell.value = translation_map[r]
            print(f"  Row {r:2d} | GRI {sheet.cell(row=r, column=4).value} | Written Translation")
        else:
            print(f"  Row {r:2d} | GRI {sheet.cell(row=r, column=4).value} | (No translation mapped)")
    else:
        print(f"  Row {r:2d} | GRI {sheet.cell(row=r, column=4).value} | Skipped (Existing value: {kpi_cell.value})")

# Step 4: Save workbook
print("Saving modified workbook...")
wb.save(file_path)
print("--- Workbook updated and saved successfully! ---")
