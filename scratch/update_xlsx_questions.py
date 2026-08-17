import os
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define paths
file_path = "/Users/congnguyen/Library/CloudStorage/OneDrive-Personal/Work/04. Company/Synodus/GOV/01.VNA/Document/Chỉ tiêu Governance_v3_Edited.xlsx"
backup_dir = "/Users/congnguyen/Library/CloudStorage/OneDrive-Personal/Work/04. Company/Synodus/GOV/01.VNA/Document/Backup"
backup_path = os.path.join(backup_dir, "Chỉ tiêu Governance_v3_Edited_Backup.xlsx")

print("--- Start Excel Update ---")

# Step 1: Create Backup
print("Checking backup directory...")
if not os.path.exists(backup_dir):
    os.makedirs(backup_dir)
    print("Created backup directory:", backup_dir)

print(f"Creating backup of original file to: {backup_path}")
shutil.copy(file_path, backup_path)
print("Backup created successfully.")

# Step 2: Load workbook
print(f"Loading workbook from: {file_path}")
wb = openpyxl.load_workbook(file_path)
sheet = wb['Chi tiết chỉ tiêu']

# Map of GRI codes to consolidated questions (1 question per indicator)
questions_map = {
    "GRI 2-9": "Cơ cấu quản trị của Tổng công ty được tổ chức như thế nào và thành phần Hội đồng quản trị (HĐQT) cùng các Ủy ban trực thuộc được phân bổ ra sao theo các tiêu chí của GRI để thực hiện vai trò định hướng và giám sát cao nhất đối với phát triển bền vững (PTBV)?",
    "GRI 2-10": "Quy trình đề cử, lựa chọn thành viên HĐQT và các tiểu ban được thực hiện như thế nào và các tiêu chí về tính độc lập, đa dạng, cùng năng lực liên quan đến PTBV được lồng ghép ra sao trong quy trình này?",
    "GRI 2-11": "Vai trò của Chủ tịch HĐQT và Tổng Giám đốc có được phân tách rõ ràng không và các biện pháp nào được áp dụng để đảm bảo tính độc lập cũng như ngăn ngừa xung đột lợi ích giữa hai chức danh này?",
    "GRI 2-12": "Vai trò của HĐQT và Ban điều hành trong việc xây dựng, phê duyệt, giám sát quy trình nhận diện tác động và định kỳ đánh giá hiệu quả các chiến lược, chính sách PTBV của Tổng công ty được quy định như thế nào?",
    "GRI 2-13": "HĐQT thực hiện phân cấp, ủy quyền trách nhiệm quản lý các tác động PTBV cho Ban điều hành cùng các Tổ công tác chuyên môn ra sao và quy trình báo cáo ngược kết quả quản lý lên HĐQT diễn ra như thế nào?",
    "GRI 2-14": "Quy trình rà soát và phê duyệt các thông tin công bố trong Báo cáo PTBV (bao gồm cả danh mục chủ đề trọng yếu) của HĐQT được thực hiện cụ thể theo các bước nào?",
    "GRI 2-15": "Quy trình ngăn ngừa, giảm thiểu và công bố thông tin liên quan đến các tình huống xung đột lợi ích (bao gồm kiêm nhiệm chéo, sở hữu chéo, cổ đông chi phối và giao dịch bên liên quan) của HĐQT được quy định ra sao?",
    "GRI 2-16": "Cơ chế tiếp nhận, xử lý và báo cáo các mối quan ngại nghiêm trọng lên HĐQT được tổ chức như thế nào và trong kỳ báo cáo có phát sinh mối quan ngại nào cần xử lý hay không?",
    "GRI 2-17": "Tổng công ty đã triển khai các chương trình đào tạo hoặc biện pháp nào để nâng cao kiến thức, kỹ năng tập thể của HĐQT về phát triển bền vững và quản trị ESG?",
    "GRI 2-18": "Quy trình, tần suất đánh giá hiệu quả hoạt động của HĐQT trong giám sát các tác động PTBV được thực hiện như thế nào (tự đánh giá hay qua tư vấn độc lập) và những cải tiến nào đã được thực hiện sau đánh giá?",
    "GRI 2-23": "Các cam kết chính sách về ứng xử kinh doanh có trách nhiệm và tôn trọng quyền con người của Tổng công ty gồm những gì, phạm vi áp dụng, cấp phê duyệt ra sao và được truyền thông công khai đến các bên liên quan bằng cách nào?",
    "GRI 2-24": "Tổng công ty thực hiện lồng ghép các cam kết chính sách vào chiến lược, hoạt động vận hành, mối quan hệ kinh doanh với nhà cung cấp ra sao và các chương trình đào tạo thực thi chính sách được triển khai như thế nào?",
    "GRI 2-25": "Tổng công ty thiết lập cơ chế khiếu nại, quy trình khắc phục các tác động tiêu cực như thế nào và làm sao để đánh giá tính hiệu quả của các cơ chế giải quyết khiếu nại này dựa trên phản hồi của các bên liên quan?",
    "GRI 2-26": "Các kênh và cơ chế để người lao động hoặc các bên liên quan ngoài Tổng công ty tìm kiếm tư vấn hoặc phản ánh các quan ngại về hành vi đạo đức kinh doanh được thiết lập như thế nào và quy trình bảo mật thông tin người phản ánh ra sao?",
    "GRI 2-27": "Quy trình kiểm soát tuân thủ pháp luật của Tổng công ty được thực hiện thế nào và trong kỳ báo cáo có ghi nhận trường hợp vi phạm đáng kể nào dẫn đến bị phạt tiền hoặc chịu chế tài phi tiền tệ không?",
    "GRI 2-29": "Tổng công ty áp dụng phương pháp nào để nhận diện, đối thoại thực chất và thu thập ý kiến phản hồi từ các nhóm bên liên quan trọng yếu về các vấn đề phát triển bền vững?",
    "GRI 3-3": "Quy trình đánh giá trọng yếu kép để xác định các chủ đề trọng yếu thuộc trụ cột Governance được thực hiện như thế nào và Tổng công ty quản lý, theo dõi tính hiệu quả của các hành động ứng phó với tác động đó ra sao?",
    "GRI 201-4": "Tổng giá trị tiền tệ các khoản trợ cấp, ưu đãi hoặc hỗ trợ tài chính mà Tổng công ty nhận được từ Chính phủ (bao gồm tiến độ của gói hỗ trợ thanh khoản 12.000 tỷ đồng) trong kỳ báo cáo là bao nhiêu và được kế toán, kiểm toán thế nào?",
    "GRI 205-2": "Tỷ lệ thành viên HĐQT, người lao động và đối tác kinh doanh được truyền thông và tham gia đào tạo về chính sách, quy trình phòng chống tham nhũng trong kỳ báo cáo đạt bao nhiêu phần trăm?",
    "GRI 205-3": "Trong kỳ báo cáo, Tổng công ty có ghi nhận sự cố tham nhũng nào được xác nhận dẫn đến kỷ luật lao động, chấm dứt hợp đồng với đối tác hoặc phát sinh vụ việc pháp lý công khai chống lại Tổng công ty không?",
    "GRI 206-1": "Tổng công ty có phát sinh vụ kiện tụng hay tranh chấp pháp lý nào (đang xử lý hoặc đã hoàn tất) liên quan đến hành vi hạn chế cạnh tranh hoặc vi phạm luật chống độc quyền trong kỳ báo cáo hay không?",
    "GRI 415-1": "Tổng công ty có thực hiện các khoản đóng góp chính trị trực tiếp hay gián tiếp (bao gồm cả phí hội viên trả cho các hiệp hội ngành nghề và liên minh hàng không) trong kỳ báo cáo hay không và giá trị cụ thể là bao nhiêu?"
}

# Helper to copy cell styles
def copy_style(src_cell, dst_cell):
    if src_cell.font:
        dst_cell.font = Font(
            name=src_cell.font.name,
            size=src_cell.font.size,
            bold=src_cell.font.bold,
            italic=src_cell.font.italic,
            charset=src_cell.font.charset,
            color=src_cell.font.color,
            underline=src_cell.font.underline,
            strike=src_cell.font.strike
        )
    if src_cell.fill:
        dst_cell.fill = PatternFill(
            fill_type=src_cell.fill.fill_type,
            start_color=src_cell.fill.start_color,
            end_color=src_cell.fill.end_color
        )
    if src_cell.border:
        dst_cell.border = Border(
            left=src_cell.border.left,
            right=src_cell.border.right,
            top=src_cell.border.top,
            bottom=src_cell.border.bottom,
            diagonal=src_cell.border.diagonal,
            diagonal_direction=src_cell.border.diagonal_direction,
            outline=src_cell.border.outline,
            vertical=src_cell.border.vertical,
            horizontal=src_cell.border.horizontal
        )
    if src_cell.alignment:
        dst_cell.alignment = Alignment(
            horizontal=src_cell.alignment.horizontal,
            vertical=src_cell.alignment.vertical,
            text_rotation=src_cell.alignment.text_rotation,
            wrap_text=src_cell.alignment.wrap_text,
            shrink_to_fit=src_cell.alignment.shrink_to_fit,
            indent=src_cell.alignment.indent
        )
    if src_cell.number_format:
        dst_cell.number_format = src_cell.number_format

# Step 3: Modify Merge Range of Row 1 (Title Header)
print("Updating title merge from A1:I1 to A1:J1...")
target_range = None
for r in list(sheet.merged_cells.ranges):
    if r.min_row == 1 and r.max_row == 1 and r.min_col == 1 and r.max_col == 9:
        target_range = r
        break

if target_range:
    sheet.unmerge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    print("Unmerged A1:I1.")
sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
print("Merged A1:J1.")

# Step 4: Write Header Cell (J3)
print("Writing header 'Câu hỏi' to J3...")
header_cell = sheet.cell(row=3, column=10)
header_cell.value = "Câu hỏi"
# Copy style from cell I3 (Chỉ tiêu bổ sung)
src_header = sheet.cell(row=3, column=9)
copy_style(src_header, header_cell)
# Ensure clean alignment
header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Step 5: Fill questions and styles for rows 4 to 25
print("Writing questions and copying styles for data rows...")
for r in range(4, sheet.max_row + 1):
    gri_cell = sheet.cell(row=r, column=2)
    gri_code = str(gri_cell.value).strip() if gri_cell.value else ""
    
    question_cell = sheet.cell(row=r, column=10)
    
    if gri_code in questions_map:
        question_cell.value = questions_map[gri_code]
        print(f"  Row {r:2d}: {gri_code} -> Written Question")
    else:
        question_cell.value = ""
        print(f"  Row {r:2d}: {gri_code} -> (No question mapped)")
        
    # Copy style from cell F{r} (Các mục chính cần hiển thị)
    src_data_cell = sheet.cell(row=r, column=6)
    copy_style(src_data_cell, question_cell)
    # Ensure text wrapping is enabled and alignment is left-top
    question_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Step 6: Set column J width
print("Setting column J width to 50...")
sheet.column_dimensions['J'].width = 50.0

# Save updated sheet
print("Saving modified workbook...")
wb.save(file_path)
print("--- Workbook updated and saved successfully! ---")
