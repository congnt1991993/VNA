import openpyxl

file_path = "/Users/congnguyen/Library/CloudStorage/OneDrive-Personal/Work/04. Company/Synodus/GOV/01.VNA/Document/Netzero_GRI Summary.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb['95 Chỉ tiêu']

print("Checking first few rows:")
print("-" * 80)
for r in range(3, 15):
    gri = sheet.cell(r, 4).value
    desc = sheet.cell(r, 7).value
    kpi_name = sheet.cell(r, 13).value
    print(f"Row {r:2d} | GRI: {gri}")
    print(f"       | G (Desc): {repr(desc)}")
    print(f"       | M (Name): {repr(kpi_name)}")
    print("-" * 50)

# Check for any remaining empty values in column M for rows with description
missing_count = 0
for r in range(3, sheet.max_row + 1):
    desc = sheet.cell(r, 7).value
    kpi_name = sheet.cell(r, 13).value
    if desc is not None and (kpi_name is None or str(kpi_name).strip() == ""):
        print(f"ERROR: Row {r} has description but is missing translation!")
        missing_count += 1

if missing_count == 0:
    print("\nSUCCESS: All rows with descriptions now have a Tên KPI translation!")
else:
    print(f"\nFAILURE: {missing_count} rows are missing a Tên KPI!")
